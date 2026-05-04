"""
genre_review.py
human_label再ラベリング補助ツール。
genre_labels_all.csv を読み込み、TAKAHIROが確認・訂正するための
Excelファイルを生成する。

出力: D:\r8_strategy\data\results\genre_review.xlsx
  - Sheet1: 要確認リスト（needs_human=True + confidence<0.5の自動確定分）
  - Sheet2: 自動確定リスト（参照用・編集不要）
"""

import os
import csv
import glob
import re
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ────────────────────────────────────────────
# 設定
# ────────────────────────────────────────────
CSV_IN      = r"D:\r8_strategy\data\results\genre_labels_all.csv"
CORPUS_DIR  = r"D:\r8_strategy\corpus\corpus_archive"
BOOK_DIR    = r"D:\r8_strategy\corpus\book"
OUT_XLSX    = r"D:\r8_strategy\data\results\genre_review.xlsx"

GENRE_NAMES = {
    "1": "投資・金融",
    "2": "カルト・宗教",
    "3": "恋愛・人間関係",
    "4": "教育・自己啓発",
    "5": "政治・陰謀",
    "6": "その他",
    1: "投資・金融",
    2: "カルト・宗教",
    3: "恋愛・人間関係",
    4: "教育・自己啓発",
    5: "政治・陰謀",
    6: "その他",
}

# book系フォルダ名 → ジャンルコード（確定）
BOOK_GENRE = {
    "kyouikunohoukai":              4,  # 教育・自己啓発
    "kimiwanazehatarakuka":         4,  # 教育・自己啓発
    "byousokude_itiokuenkasegujouken": 1,  # 投資・金融
    "butinukutikara":               4,  # 教育・自己啓発
}

# ────────────────────────────────────────────
# book系パス解決
# ────────────────────────────────────────────
def resolve_book_genre(target):
    """
    target が book パスの場合、フォルダ名からジャンルを確定する。
    例: 'corpus\book\kyouikunohoukai\ch1.txt' → (4, 1.0, 'book_folder_rule')
    """
    # CSVのtarget列に書かれているパス形式を正規化
    normalized = target.replace("\\", "/").lower()
    for folder, code in BOOK_GENRE.items():
        if folder in normalized:
            return code, 1.0, "book_folder_rule"
    # bookディレクトリだがフォルダ未登録
    if "book" in normalized:
        return 6, 0.0, "book_unknown"
    return None

# ────────────────────────────────────────────
# テキスト冒頭抽出（確認用スニペット）
# ────────────────────────────────────────────
def get_snippet(txt_path, chars=120):
    """テキストの先頭を返す。URLのみの行はスキップ。"""
    if not txt_path or txt_path == "NOT_FOUND":
        return ""
    # book系のパス解決
    if not os.path.exists(txt_path):
        # corpus\book\... 形式のパスを絶対パスに変換
        alt = txt_path.replace("corpus\\", r"D:\r8_strategy\corpus\\")
        alt = alt.replace("corpus/", r"D:\r8_strategy\corpus/")
        if os.path.exists(alt):
            txt_path = alt
        else:
            return "（txtファイル未取得）"
    try:
        with open(txt_path, encoding="utf-8", errors="ignore") as f:
            lines = f.readlines()
        # URL行・空行をスキップして本文を取得
        text_lines = []
        for line in lines:
            line = line.strip()
            if not line:
                continue
            if line.startswith("http") and len(line.split()) == 1:
                continue
            if line.startswith("[SOURCE]") or line.startswith("[DATE]") or \
               line.startswith("[CATEGORY]") or line.startswith("[TEXT]"):
                continue
            text_lines.append(line)
        snippet = "　".join(text_lines[:3])
        return snippet[:chars]
    except Exception as e:
        return f"読み取りエラー: {e}"

# ────────────────────────────────────────────
# CSV読み込み・分類
# ────────────────────────────────────────────
def load_and_classify(csv_path):
    with open(csv_path, encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        rows = list(reader)

    review_rows = []    # 要確認
    confirmed_rows = [] # 自動確定

    for row in rows:
        target = row["target"].strip()
        genre_code = row["genre_code"]
        genre_name = row["genre_name"]
        confidence = float(row["confidence"])
        needs_human = row["needs_human"].strip().lower() == "true"
        reason = row["reason"]
        human_label = row["human_label"]
        txt_path = row["txt_path"]

        # book系はフォルダルールで上書き
        book_result = resolve_book_genre(target)
        if book_result:
            code, conf, rsn = book_result
            if rsn == "book_folder_rule":
                genre_code = str(code)
                genre_name = GENRE_NAMES[code]
                confidence = conf
                needs_human = False
                reason = rsn

        # txt_not_found（book以外）は要確認
        if reason == "txt_not_found" and not (book_result and book_result[2] == "book_folder_rule"):
            needs_human = True

        # confidence<0.5の自動確定分を要確認に移動
        if not needs_human and confidence < 0.5:
            needs_human = True

        snippet = get_snippet(txt_path)

        record = {
            "target": target,
            "human_label": human_label,
            "auto_genre_code": genre_code,
            "auto_genre_name": genre_name,
            "confidence": round(confidence, 3),
            "reason": reason,
            "snippet": snippet,
        }

        if needs_human:
            review_rows.append(record)
        else:
            confirmed_rows.append(record)

    return review_rows, confirmed_rows

# ────────────────────────────────────────────
# Excel生成
# ────────────────────────────────────────────
def make_excel(review_rows, confirmed_rows, out_path):
    wb = openpyxl.Workbook()

    # ── Sheet1: 要確認リスト ──────────────────
    ws1 = wb.active
    ws1.title = "要確認（TAKAHIRO記入）"

    # ヘッダー
    headers = [
        "target",
        "CMI human_label",
        "推定ジャンル",
        "信頼度",
        "【記入】genre_code",
        "【記入】genre_name",
        "テキスト冒頭",
        "reason",
    ]

    # スタイル定義
    header_fill = PatternFill("solid", fgColor="1F497D")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    input_fill  = PatternFill("solid", fgColor="FFFFD0")  # 薄黄：記入欄
    auto_fill   = PatternFill("solid", fgColor="E8F0FE")  # 薄青：自動推定
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    center = Alignment(horizontal="center", vertical="center", wrap_text=False)
    wrap   = Alignment(horizontal="left",   vertical="top",    wrap_text=True)

    # ヘッダー行
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    # データ行
    for row_i, rec in enumerate(review_rows, 2):
        vals = [
            rec["target"],
            rec["human_label"],
            f"{rec['auto_genre_code']}: {rec['auto_genre_name']}",
            rec["confidence"],
            "",   # 記入欄: genre_code
            "",   # 記入欄: genre_name
            rec["snippet"],
            rec["reason"],
        ]
        for col, val in enumerate(vals, 1):
            cell = ws1.cell(row=row_i, column=col, value=val)
            cell.border = border
            cell.font = Font(name="Arial", size=10)
            if col in (5, 6):  # 記入欄
                cell.fill = input_fill
                cell.alignment = center
            elif col == 7:     # スニペット
                cell.alignment = wrap
            else:
                cell.alignment = center

    # 凡例をA1コメントとして追加
    ws1["E1"].comment = None  # 既存コメントクリア

    # 凡例行（最終行+2）
    legend_row = len(review_rows) + 3
    ws1.cell(row=legend_row, column=1, value="【ジャンルコード凡例】").font = Font(bold=True)
    for i, (code, name) in enumerate([(1,"投資・金融"),(2,"カルト・宗教"),(3,"恋愛・人間関係"),
                                       (4,"教育・自己啓発"),(5,"政治・陰謀"),(6,"その他")], 1):
        ws1.cell(row=legend_row+i, column=1, value=f"{code} = {name}")

    # 列幅
    col_widths = [14, 14, 22, 10, 14, 18, 45, 18]
    for col, w in enumerate(col_widths, 1):
        ws1.column_dimensions[get_column_letter(col)].width = w

    # 行高（スニペット列）
    for row_i in range(2, len(review_rows) + 2):
        ws1.row_dimensions[row_i].height = 32

    # ウィンドウ枠固定
    ws1.freeze_panes = "A2"

    # ── Sheet2: 自動確定リスト（参照用）────────
    ws2 = wb.create_sheet("自動確定（参照用）")
    headers2 = ["target", "CMI human_label", "確定ジャンル", "信頼度", "reason"]
    confirmed_header_fill = PatternFill("solid", fgColor="375623")
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.fill = confirmed_header_fill
        cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        cell.alignment = center
        cell.border = border

    for row_i, rec in enumerate(confirmed_rows, 2):
        vals = [
            rec["target"],
            rec["human_label"],
            f"{rec['auto_genre_code']}: {rec['auto_genre_name']}",
            rec["confidence"],
            rec["reason"],
        ]
        for col, val in enumerate(vals, 1):
            cell = ws2.cell(row=row_i, column=col, value=val)
            cell.border = border
            cell.font = Font(name="Arial", size=10)
            cell.alignment = center

    col_widths2 = [14, 14, 22, 10, 18]
    for col, w in enumerate(col_widths2, 1):
        ws2.column_dimensions[get_column_letter(col)].width = w
    ws2.freeze_panes = "A2"

    wb.save(out_path)
    print(f"出力完了: {out_path}")
    print(f"  Sheet1 要確認: {len(review_rows)}件")
    print(f"  Sheet2 自動確定: {len(confirmed_rows)}件")

# ────────────────────────────────────────────
# 実行
# ────────────────────────────────────────────
if __name__ == "__main__":
    print("読み込み中...")
    review_rows, confirmed_rows = load_and_classify(CSV_IN)
    print(f"要確認: {len(review_rows)}件 / 自動確定: {len(confirmed_rows)}件")
    make_excel(review_rows, confirmed_rows, OUT_XLSX)
