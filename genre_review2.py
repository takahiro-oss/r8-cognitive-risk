"""
genre_review.py 修正版
エスケープ修正済み・openpyxl自動インストール対応
"""

import os
import csv
import subprocess
import sys

try:
    import openpyxl
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl

from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

CSV_IN   = r"D:\r8_strategy\data\results\genre_labels_all.csv"
OUT_XLSX = r"D:\r8_strategy\data\results\genre_review.xlsx"

GENRE_NAMES = {
    "1": "投資・金融",   1: "投資・金融",
    "2": "カルト・宗教", 2: "カルト・宗教",
    "3": "恋愛・人間関係", 3: "恋愛・人間関係",
    "4": "教育・自己啓発", 4: "教育・自己啓発",
    "5": "政治・陰謀",   5: "政治・陰謀",
    "6": "その他",       6: "その他",
}

BOOK_GENRE = {
    "kyouikunohoukai":                 4,
    "kimiwanazehatarakuka":            4,
    "byousokude_itiokuenkasegujouken": 1,
    "butinukutikara":                  4,
}

def resolve_book_genre(target):
    normalized = target.replace("\\", "/").lower()
    for folder, code in BOOK_GENRE.items():
        if folder in normalized:
            return code, 1.0, "book_folder_rule"
    if "book" in normalized:
        return 6, 0.0, "book_unknown"
    return None

def get_snippet(txt_path, chars=120):
    if not txt_path or txt_path == "NOT_FOUND":
        return ""
    if not os.path.exists(txt_path):
        return "(txtファイル未取得)"
    try:
        with open(txt_path, encoding="utf-8", errors="ignore") as f:
            lines = f.readlines()
        text_lines = []
        for line in lines:
            line = line.strip()
            if not line:
                continue
            if line.startswith("http") and len(line.split()) == 1:
                continue
            if line.startswith(("[SOURCE]", "[DATE]", "[CATEGORY]", "[TEXT]")):
                continue
            text_lines.append(line)
        return ("  ".join(text_lines[:3]))[:chars]
    except Exception as e:
        return f"エラー: {e}"

def load_and_classify(csv_path):
    with open(csv_path, encoding="utf-8-sig") as f:
        rows = list(csv.DictReader(f))

    review_rows    = []
    confirmed_rows = []

    for row in rows:
        target      = row["target"].strip()
        genre_code  = row["genre_code"]
        genre_name  = row["genre_name"]
        confidence  = float(row["confidence"])
        needs_human = row["needs_human"].strip().lower() == "true"
        reason      = row["reason"]
        human_label = row["human_label"]
        txt_path    = row["txt_path"]

        book_result = resolve_book_genre(target)
        if book_result:
            code, conf, rsn = book_result
            if rsn == "book_folder_rule":
                genre_code  = str(code)
                genre_name  = GENRE_NAMES[code]
                confidence  = conf
                needs_human = False
                reason      = rsn

        if reason == "txt_not_found" and not (book_result and book_result[2] == "book_folder_rule"):
            needs_human = True

        if not needs_human and confidence < 0.5:
            needs_human = True

        record = {
            "target":          target,
            "human_label":     human_label,
            "auto_genre_code": genre_code,
            "auto_genre_name": genre_name,
            "confidence":      round(confidence, 3),
            "reason":          reason,
            "snippet":         get_snippet(txt_path),
        }

        (review_rows if needs_human else confirmed_rows).append(record)

    return review_rows, confirmed_rows

def make_excel(review_rows, confirmed_rows, out_path):
    wb = openpyxl.Workbook()
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin")
    )
    center = Alignment(horizontal="center", vertical="center")
    wrap   = Alignment(horizontal="left",   vertical="top", wrap_text=True)

    # Sheet1
    ws1 = wb.active
    ws1.title = "要確認（TAKAHIRO記入）"
    headers1 = [
        "target", "CMI human_label", "推定ジャンル", "信頼度",
        "【記入】genre_code", "【記入】genre_name", "テキスト冒頭", "reason",
    ]
    hdr_fill = PatternFill("solid", fgColor="1F497D")
    hdr_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    inp_fill = PatternFill("solid", fgColor="FFFFD0")

    for col, h in enumerate(headers1, 1):
        c = ws1.cell(row=1, column=col, value=h)
        c.fill, c.font, c.alignment, c.border = hdr_fill, hdr_font, center, border

    for ri, rec in enumerate(review_rows, 2):
        vals = [
            rec["target"], rec["human_label"],
            f"{rec['auto_genre_code']}: {rec['auto_genre_name']}",
            rec["confidence"], "", "",
            rec["snippet"], rec["reason"],
        ]
        for col, val in enumerate(vals, 1):
            c = ws1.cell(row=ri, column=col, value=val)
            c.border = border
            c.font   = Font(name="Arial", size=10)
            if col in (5, 6):
                c.fill      = inp_fill
                c.alignment = center
            elif col == 7:
                c.alignment = wrap
            else:
                c.alignment = center

    lr = len(review_rows) + 3
    ws1.cell(row=lr, column=1, value="【ジャンルコード凡例】").font = Font(bold=True, name="Arial")
    for i, (code, name) in enumerate(
        [(1,"投資・金融"),(2,"カルト・宗教"),(3,"恋愛・人間関係"),
         (4,"教育・自己啓発"),(5,"政治・陰謀"),(6,"その他")], 1
    ):
        ws1.cell(row=lr+i, column=1, value=f"{code} = {name}").font = Font(name="Arial", size=10)

    for col, w in enumerate([14,14,22,10,14,18,45,18], 1):
        ws1.column_dimensions[get_column_letter(col)].width = w
    for ri in range(2, len(review_rows)+2):
        ws1.row_dimensions[ri].height = 32
    ws1.freeze_panes = "A2"

    # Sheet2
    ws2 = wb.create_sheet("自動確定（参照用）")
    headers2 = ["target", "CMI human_label", "確定ジャンル", "信頼度", "reason"]
    hdr_fill2 = PatternFill("solid", fgColor="375623")
    for col, h in enumerate(headers2, 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.fill = hdr_fill2
        c.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.alignment, c.border = center, border
    for ri, rec in enumerate(confirmed_rows, 2):
        vals = [
            rec["target"], rec["human_label"],
            f"{rec['auto_genre_code']}: {rec['auto_genre_name']}",
            rec["confidence"], rec["reason"],
        ]
        for col, val in enumerate(vals, 1):
            c = ws2.cell(row=ri, column=col, value=val)
            c.border, c.font, c.alignment = border, Font(name="Arial", size=10), center
    for col, w in enumerate([14,14,22,10,18], 1):
        ws2.column_dimensions[get_column_letter(col)].width = w
    ws2.freeze_panes = "A2"

    wb.save(out_path)
    print(f"出力完了: {out_path}")
    print(f"  Sheet1 要確認:   {len(review_rows)}件")
    print(f"  Sheet2 自動確定: {len(confirmed_rows)}件")

if __name__ == "__main__":
    print("読み込み中...")
    review_rows, confirmed_rows = load_and_classify(CSV_IN)
    print(f"要確認: {len(review_rows)}件 / 自動確定: {len(confirmed_rows)}件")
    make_excel(review_rows, confirmed_rows, OUT_XLSX)
