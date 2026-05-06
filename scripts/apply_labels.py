# apply_labels.py
# audit_results_v4.xlsm の human_label / riskfactor / remarks を
# corpus_master.csv に反映するスクリプト
#
# 使用方法:
#   .\.venv\Scripts\python.exe scripts\apply_labels.py --dry-run
#   .\.venv\Scripts\python.exe scripts\apply_labels.py
#
# 実行条件: human_label再ラベリング完了後に実行すること

import csv
import sys
import shutil
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("[ERROR] openpyxl が見つかりません。pip install openpyxl を実行してください。")
    sys.exit(1)

# ── パス設定 ──────────────────────────────────────────────
BASE_DIR  = Path(r"D:\r8_strategy")
XLSM_PATH = BASE_DIR / "data" / "results" / "audit_results_v4.xlsm"
CSV_PATH  = BASE_DIR / "data" / "results" / "corpus_master.csv"

# ── xlsm列番号（1-indexed）──────────────────────────────
COL_TARGET  = 2
COL_HLABEL  = 20
COL_RF1     = 21
COL_RF2     = 22
COL_RF3     = 23
COL_REMARKS = 24

# Apple私用領域文字（Mac由来のxlsmに混入する不可視文字）
APPLE_PUA_CHARS = "\uF8F0\uF8F1\uF8F2\uF8F3\uF8F4\uF8FF"


def strip_apple_pua(s: str) -> str:
    """Apple私用領域文字（U+F8F0等）を除去する"""
    for c in APPLE_PUA_CHARS:
        s = s.replace(c, "")
    return s


def normalize_target(raw: str) -> str:
    """
    xlsm内のtarget文字列をcorpus_masterの識別子形式に正規化する。
    """
    if raw is None:
        return ""
    s = strip_apple_pua(str(raw).strip().replace("/", "\\"))

    # フルパス（ドライブレター付き: D:\...）
    if len(s) >= 3 and s[1] == ":" and s[2] == "\\":
        return Path(s).stem

    # corpus\book\{folder}\{file}
    if s.startswith("corpus\\book\\"):
        parts = s.split("\\")
        if len(parts) >= 4:
            folder = parts[2]
            stem   = Path(parts[3]).stem
            return f"book_{folder}_{stem}"

    # corpus\phase2\scan\{file}
    if s.startswith("corpus\\phase2\\"):
        return Path(s.split("\\")[-1]).stem

    # corpus\corpus_archive\{file}
    if s.startswith("corpus\\corpus_archive\\"):
        return Path(s.split("\\")[-1]).stem

    # 拡張子除去（ir_report.pdf → ir_report）
    p = Path(s)
    if p.suffix in (".pdf", ".txt", ".csv"):
        return p.stem

    return s


def load_xlsm(path: Path) -> dict:
    """
    xlsm から {normalized_target: {human_label, rf1, rf2, rf3, remarks}} を返す。
    human_label が空の行はスキップ。
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["all_corpus"]

    records = {}
    skipped_empty = 0

    for row in range(2, ws.max_row + 1):
        raw_target = ws.cell(row, COL_TARGET).value
        if not raw_target:
            continue

        hl  = ws.cell(row, COL_HLABEL).value
        rf1 = ws.cell(row, COL_RF1).value
        rf2 = ws.cell(row, COL_RF2).value
        rf3 = ws.cell(row, COL_RF3).value
        rem = ws.cell(row, COL_REMARKS).value

        if not hl:
            skipped_empty += 1
            continue

        key = normalize_target(str(raw_target))
        records[key] = {
            "human_label":  str(hl).strip(),
            "riskfactor_1": str(rf1).strip() if rf1 else "",
            "riskfactor_2": str(rf2).strip() if rf2 else "",
            "riskfactor_3": str(rf3).strip() if rf3 else "",
            "remarks":      str(rem).strip() if rem else "",
        }

    if skipped_empty:
        print(f"[INFO] human_label空欄のためスキップした行: {skipped_empty}件")

    return records


def apply_labels(dry_run: bool = False):
    print(f"[START] apply_labels.py  dry_run={dry_run}")
    print(f"  xlsm: {XLSM_PATH}")
    print(f"  csv:  {CSV_PATH}")

    print("\n[1/4] xlsm 読み込み中...")
    xlsm_records = load_xlsm(XLSM_PATH)
    print(f"  取得レコード数: {len(xlsm_records)}")

    print("\n[2/4] corpus_master.csv 読み込み中...")
    with open(CSV_PATH, encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        fieldnames = reader.fieldnames
        rows = list(reader)
    print(f"  行数: {len(rows)}")

    print("\n[3/4] 照合中...")
    csv_targets = {row["target"].strip() for row in rows}
    not_in_csv  = [k for k in xlsm_records if k not in csv_targets]
    not_in_xlsm = [t for t in csv_targets  if t not in xlsm_records]

    if not_in_csv:
        print(f"\n[ERROR] xlsmにあるがcorpus_masterに存在しないtarget: {len(not_in_csv)}件")
        for t in sorted(not_in_csv):
            print(f"  x {t}  (hex: {t.encode('utf-8').hex()})")
        print("\n[ABORT] 書き込みを中止します。")
        return

    if not_in_xlsm:
        print(f"[WARN] corpus_masterにあるがxlsmにないtarget: {len(not_in_xlsm)}件（上書きしない）")
        for t in sorted(not_in_xlsm)[:10]:
            print(f"  - {t}")
        if len(not_in_xlsm) > 10:
            print(f"  ... 他 {len(not_in_xlsm)-10}件")

    updated_targets = []
    diff_details    = []
    for row in rows:
        t = row["target"].strip()
        if t not in xlsm_records:
            continue
        new = xlsm_records[t]
        changed = []
        for field in ["human_label", "riskfactor_1", "riskfactor_2", "riskfactor_3", "remarks"]:
            old_val = row.get(field, "").strip()
            new_val = new.get(field, "").strip()
            if old_val != new_val:
                changed.append(f"{field}: [{old_val}] -> [{new_val}]")
        if changed:
            updated_targets.append(t)
            diff_details.append((t, changed))

    print(f"\n  変更あり: {len(updated_targets)}件 / 変更なし: {len(xlsm_records) - len(updated_targets)}件")

    if dry_run:
        print("\n[DRY-RUN] 差分プレビュー（書き込みは行いません）:")
        for t, fields in diff_details[:20]:
            print(f"  {t}")
            for f in fields:
                print(f"    {f}")
        if len(diff_details) > 20:
            print(f"  ... 他 {len(diff_details)-20}件")
        print("\n[DRY-RUN] 完了。--dry-run なしで実行すると書き込みます。")
        return

    if not updated_targets:
        print("[INFO] 差分なし。書き込みをスキップします。")
        return

    print("\n[4/4] 書き込み中...")
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = CSV_PATH.with_name(f"corpus_master_backup_{ts}.csv")
    shutil.copy2(CSV_PATH, backup_path)
    print(f"  バックアップ作成: {backup_path.name}")

    for row in rows:
        t = row["target"].strip()
        if t not in xlsm_records:
            continue
        new = xlsm_records[t]
        row["human_label"]  = new["human_label"]
        row["riskfactor_1"] = new["riskfactor_1"]
        row["riskfactor_2"] = new["riskfactor_2"]
        row["riskfactor_3"] = new["riskfactor_3"]
        row["remarks"]      = new["remarks"]

    with open(CSV_PATH, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)

    print(f"  corpus_master.csv 更新完了: {len(updated_targets)}件")
    print(f"\n[DONE]")


if __name__ == "__main__":
    dry_run = "--dry-run" in sys.argv
    apply_labels(dry_run=dry_run)
