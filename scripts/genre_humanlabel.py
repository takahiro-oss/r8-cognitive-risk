"""
genre_humanlabel.py
human_label再ラベリング + ジャンル確定を同時に行うExcel作成ツール。

【記入ルール】
- F列（human_label訂正）: 空欄=現在値維持 / H・M・Lを入力=変更
- G列（ジャンル訂正）:     空欄=推定値採用 / 1〜6を入力=変更

出力: D:/r8_strategy/data/results/humanlabel_review.xlsx
"""

import os, csv, sys, subprocess

try:
    import openpyxl
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl

from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ────────────────────────────────────────────
# パス設定
# ────────────────────────────────────────────
AUDIT_CSV   = r"D:\r8_strategy\data\results\audit_results_v3.csv"
GENRE_CSV   = r"D:\r8_strategy\data\results\genre_labels_all.csv"
OUT_XLSX    = r"D:\r8_strategy\data\results\humanlabel_review.xlsx"

GENRE_NAMES = {
    "1": "投資・金融",   1: "投資・金融",
    "2": "カルト・宗教", 2: "カルト・宗教",
    "3": "恋愛・人間関係", 3: "恋愛・人間関係",
    "4": "教育・自己啓発", 4: "教育・自己啓発",
    "5": "政治・陰謀",   5: "政治・陰謀",
    "6": "その他",       6: "その他",
}

BOOK_GENRE = {
    "kyouikunohoukai":                 4,
    "kimiwanazehatarakuka":            4,
    "byousokude_itiokuenkasegujouken": 1,
    "butinukutikara":                  4,
}

# ────────────────────────────────────────────
# ユーティリティ
# ────────────────────────────────────────────
def resolve_book_genre(target):
    normalized = target.replace("\\", "/").lower()
    for folder, code in BOOK_GENRE.items():
        if folder in normalized:
            return str(code), GENRE_NAMES[code]
    return None, None

def get_auto_genre(genre_row):
    """genre_labels_all.csv の1行からジャンルコード・名称を返す（book補正込み）"""
    target     = genre_row["target"].strip()
    code       = genre_row["genre_code"]
    name       = genre_row["genre_name"]
    confidence = float(genre_row["confidence"])
    reason     = genre_row["reason"]

    # book系はフォルダルールで確定
    b_code, b_name = resolve_book_genre(target)
    if b_code:
        return b_code, b_name, 1.0, "book_folder_rule"

    return code, name, confidence, reason

def get_snippet(txt_path, chars=150):
    if not txt_path or txt_path == "NOT_FOUND":
        return ""
    if not os.path.exists(txt_path):
        return "(txtファイル未取得)"
    try:
        with open(txt_path, encoding="utf-8", errors="ignore") as f:
            lines = f.readlines()
        out = []
        for line in lines:
            line = line.strip()
            if not line:
                continue
            if line.startswith("http") and len(line.split()) == 1:
                continue
            if line.startswith(("[SOURCE]","[DATE]","[CATEGORY]","[TEXT]")):
                continue
            out.append(line)
        return ("  ".join(out[:4]))[:chars]
    except Exception as e:
        return f"エラー:{e}"

# ────────────────────────────────────────────
# データ読み込み・結合
# ────────────────────────────────────────────
def load_data():
    # audit_results_v3.csv → target: {cmi, human_label}
    audit = {}
    with open(AUDIT_CSV, encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            t = row["target"].strip()
            audit[t] = {
                "cmi":         row.get("cmi", ""),
                "human_label": row.get("human_label", "").strip(),
                "level":       row.get("level", "").strip(),
            }

    # genre_labels_all.csv → target: ジャンル情報
    genre = {}
    with open(GENRE_CSV, encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            t = row["target"].strip()
            code, name, conf, reason = get_auto_genre(row)
            genre[t] = {
                "auto_genre_code": code,
                "auto_genre_name": name,
                "confidence":      round(conf, 3),
                "reason":          reason,
                "txt_path":        row.get("txt_path", ""),
            }

    # 結合
    rows = []
    for target, a in audit.items():
        g = genre.get(target, {
            "auto_genre_code": "6",
            "auto_genre_name": "その他",
            "confidence": 0.0,
            "reason": "not_in_genre_csv",
            "txt_path": "",
        })
        rows.append({
            "target":          target,
            "cmi":             a["cmi"],
            "human_label":     a["human_label"],
            "level":           a["level"],
            "auto_genre_code": g["auto_genre_code"],
            "auto_genre_name": g["auto_genre_name"],
            "confidence":      g["confidence"],
            "reason":          g["reason"],
            "snippet":         get_snippet(g["txt_path"]),
        })

    return rows

# ────────────────────────────────────────────
# Excel生成
# ────────────────────────────────────────────
def make_excel(rows, out_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ラベリング作業シート"

    # スタイル
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin")
    )
    center  = Alignment(horizontal="center", vertical="center")
    wrap    = Alignment(horizontal="left",   vertical="top", wrap_text=True)

    # 色定義
    fill_header   = PatternFill("solid", fgColor="1F497D")  # 紺：ヘッダー
    fill_input_hl = PatternFill("solid", fgColor="FFF2CC")  # 薄黄：human_label記入欄
    fill_input_gn = PatternFill("solid", fgColor="E2EFDA")  # 薄緑：ジャンル記入欄
    fill_ref      = PatternFill("solid", fgColor="F2F2F2")  # 薄灰：参照列

    font_hdr  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    font_body = Font(name="Arial", size=10)
    font_bold = Font(name="Arial", bold=True, size=10)

    # ── ヘッダー ─────────────────────────────
    headers = [
        "target",           # A
        "CMI",              # B
        "現human_label",    # C  ← 現在値（参照）
        "R8 level",         # D  ← R8自動判定（参照）
        "推定ジャンル",     # E  ← 参照
        "信頼度",           # F  ← 参照
        "テキスト冒頭",     # G  ← 参照
        "【記入】\nhuman_label訂正",  # H ← 記入欄
        "【記入】\nジャンル訂正",     # I ← 記入欄
    ]

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill      = fill_header
        c.font      = font_hdr
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = border
    ws.row_dimensions[1].height = 30

    # ── データ行 ─────────────────────────────
    for ri, rec in enumerate(rows, 2):
        genre_disp = f"{rec['auto_genre_code']}: {rec['auto_genre_name']}"
        conf_disp  = rec["confidence"] if rec["confidence"] > 0 else "—"

        vals = [
            rec["target"],       # A
            rec["cmi"],          # B
            rec["human_label"],  # C
            rec["level"],        # D
            genre_disp,          # E
            conf_disp,           # F
            rec["snippet"],      # G
            "",                  # H 記入欄
            "",                  # I 記入欄
        ]

        for col, val in enumerate(vals, 1):
            c = ws.cell(row=ri, column=col, value=val)
            c.border = border
            c.font   = font_body

            if col == 8:   # H: human_label記入欄
                c.fill      = fill_input_hl
                c.alignment = center
            elif col == 9: # I: ジャンル記入欄
                c.fill      = fill_input_gn
                c.alignment = center
            elif col == 7: # G: テキスト冒頭
                c.fill      = fill_ref
                c.alignment = wrap
            elif col in (3, 4): # C,D: 現在値参照
                c.fill      = fill_ref
                c.alignment = center
            else:
                c.alignment = center

        ws.row_dimensions[ri].height = 36

    # ── 凡例行 ───────────────────────────────
    lr = len(rows) + 3

    # human_label凡例
    ws.cell(row=lr, column=1, value="【H列記入ルール】").font = font_bold
    ws.cell(row=lr+1, column=1, value="空欄 = 現在値を維持").font = font_body
    ws.cell(row=lr+2, column=1, value="H / M / L を入力 = 変更").font = font_body

    # ジャンル凡例
    ws.cell(row=lr, column=4, value="【I列 ジャンルコード凡例】").font = font_bold
    legends = [(1,"投資・金融"),(2,"カルト・宗教"),(3,"恋愛・人間関係"),
               (4,"教育・自己啓発"),(5,"政治・陰謀"),(6,"その他")]
    for i, (code, name) in enumerate(legends, 1):
        ws.cell(row=lr+i, column=4, value=f"{code} = {name}").font = font_body

    ws.cell(row=lr+len(legends)+2, column=4,
            value="空欄 = 推定値をそのまま採用").font = font_body

    # ── 列幅・固定 ───────────────────────────
    col_widths = [14, 8, 14, 10, 22, 8, 48, 16, 14]
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = "A2"

    wb.save(out_path)
    print(f"出力完了: {out_path}")
    print(f"  総件数: {len(rows)}件")

# ────────────────────────────────────────────
if __name__ == "__main__":
    print("読み込み中...")
    rows = load_data()
    print(f"結合完了: {len(rows)}件")
    make_excel(rows, OUT_XLSX)
    print("完了。")
