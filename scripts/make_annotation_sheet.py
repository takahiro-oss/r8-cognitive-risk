"""
make_annotation_sheet.py
アノテーター向けExcel記入シートを生成するスクリプト
出力先: D:\r8_strategy\annotator_corpus\R8_annotation_sheet.xlsx
実行: .\.venv\Scripts\python.exe scripts\make_annotation_sheet.py
"""

import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from pathlib import Path

BASE       = Path(r"D:\r8_strategy")
MASTER_CSV = BASE / "data" / "results" / "corpus_master.csv"
OUT_PATH   = BASE / "annotator_corpus" / "R8_annotation_sheet.xlsx"

with open(MASTER_CSV, encoding="utf-8-sig") as f:
    rows = list(csv.DictReader(f))

targets_ja = [r for r in rows if r.get("is_english","0")=="0" and float(r.get("cmi","0"))>0]
targets_en = [r for r in rows if r.get("is_english","0")=="1" and float(r.get("cmi","0"))>0]
targets_ja.sort(key=lambda r: float(r["cmi"]), reverse=True)
targets_en.sort(key=lambda r: float(r["cmi"]), reverse=True)
all_targets = targets_ja + targets_en
print(f"日本語: {len(targets_ja)}件 / 英語翻訳: {len(targets_en)}件 / 合計: {len(all_targets)}件")

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "判定シート"

header_font  = Font(name="Arial", bold=True, size=11, color="FFFFFF")
header_fill  = PatternFill(fill_type="solid", fgColor="2F4F8F")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
data_font    = Font(name="Arial", size=10)
data_center  = Alignment(horizontal="center", vertical="center")
data_left    = Alignment(horizontal="left",   vertical="center", wrap_text=True)
thin = Border(left=Side(style="thin"), right=Side(style="thin"),
              top=Side(style="thin"),  bottom=Side(style="thin"))

genre_fills = {
    "1": PatternFill(fill_type="solid", fgColor="FFF2CC"),
    "2": PatternFill(fill_type="solid", fgColor="FCE4D6"),
    "3": PatternFill(fill_type="solid", fgColor="E2EFDA"),
    "4": PatternFill(fill_type="solid", fgColor="DDEBF7"),
    "5": PatternFill(fill_type="solid", fgColor="EAD1DC"),
    "6": PatternFill(fill_type="solid", fgColor="F2F2F2"),
}

headers = [
    ("No.", 5), ("対象ID", 32), ("ジャンル", 18), ("CMI", 8),
    ("R8判定", 10), ("TAKAHIROラベル", 14),
    ("あなたの判定\nHIGH/MEDIUM/LOW/不明", 22), ("備考・判定理由（任意）", 35),
]
for ci, (h, w) in enumerate(headers, 1):
    cell = ws.cell(row=1, column=ci, value=h)
    cell.font, cell.fill, cell.alignment, cell.border = header_font, header_fill, header_align, thin
    ws.column_dimensions[get_column_letter(ci)].width = w
ws.row_dimensions[1].height = 40
ws.freeze_panes = "A2"

dv = DataValidation(type="list", formula1='"HIGH,MEDIUM,LOW,不明"', allow_blank=True,
                    showErrorMessage=True, errorTitle="入力エラー",
                    error="HIGH / MEDIUM / LOW / 不明 のいずれかを選択してください")
ws.add_data_validation(dv)

genre_short = {"1":"投資・金融","2":"カルト・宗教","3":"恋愛・人間関係",
               "4":"教育・自己啓発","5":"政治・陰謀","6":"その他"}

def genre_code(s):
    return s[0] if s and s[0].isdigit() else "6"

for ri, r in enumerate(all_targets, 2):
    target = r["target"]
    gc     = genre_code(r.get("genre_label","6"))
    cmi    = round(float(r.get("cmi",0)), 1)
    level  = r.get("level","")
    hlabel = r.get("human_label","")
    is_en  = r.get("is_english","0")=="1"
    gdisp  = genre_short.get(gc,"その他") + ("（英→日翻訳）" if is_en else "")
    fill   = genre_fills.get(gc, genre_fills["6"])

    for ci, val in enumerate([ri-1, target, gdisp, cmi, level, hlabel, "", ""], 1):
        c = ws.cell(row=ri, column=ci, value=val)
        c.font, c.fill, c.border = data_font, fill, thin
        c.alignment = data_center if ci in (1,4,5,6,7) else data_left
    dv.add(ws.cell(row=ri, column=7))
    ws.row_dimensions[ri].height = 18

ws2 = wb.create_sheet("使い方・判定基準概要")
gf = Font(name="Arial", size=11)
tf = Font(name="Arial", size=13, bold=True)
guide = [
    ("R8コーパス アノテーション記入シート", True),
    ("", False),
    ("【使い方】", True),
    ("1. 「判定シート」タブを開いてください", False),
    ("2. 各テキストファイルを annotator_corpus/ フォルダから開いて読んでください", False),
    ("   ・日本語テキスト: {対象ID}.txt", False),
    ("   ・英語→日本語翻訳: {対象ID}_ja.txt（ジャンル列に「英→日翻訳」と記載）", False),
    ("3. G列「あなたの判定」のプルダウンから HIGH / MEDIUM / LOW / 不明 を選んでください", False),
    ("4. 迷った場合は H列「備考」に理由を記入してください（任意）", False),
    ("", False),
    ("【判定基準の概要】", True),
    ("HIGH   : 認知操作リスクが高い。複数の操作的シグナルが確認される", False),
    ("MEDIUM : 認知操作的な要素はあるが、HIGH ほどではない", False),
    ("LOW    : 操作的シグナルがほとんど見られない", False),
    ("不明   : 判断できない（テキストが短すぎる・文脈不足等）", False),
    ("", False),
    ("【重要な注意事項】", True),
    ("・テキストの主張が正しいかどうかではなく、言語構造を評価してください", False),
    ("・「怪しそう」「内容が嫌い」という印象ではなく、判定基準に沿って判断してください", False),
    ("・テキスト中の【補完:〇〇】は文字欠損の復元箇所です（判定には影響しません）", False),
    ("・末尾にサイトナビ等が混入している場合があります。本文部分のみを評価してください", False),
    ("", False),
    ("詳細な判定基準は添付の annotation_criteria_v0.8.1.md を参照してください", False),
    ("所要時間の目安: 約20時間（1件2〜5分 × 207件）", False),
]
for i, (text, is_title) in enumerate(guide, 1):
    c = ws2.cell(row=i, column=1, value=text)
    c.font = tf if is_title else gf
    c.alignment = Alignment(horizontal="left", vertical="center")
ws2.column_dimensions["A"].width = 80

wb.save(OUT_PATH)
print(f"保存完了: {OUT_PATH}")
