"""
migrate_to_master.py
既存データをcorpus_master.csv + corpus_paths.csvに移行する。

入力:
  data/results/audit_results_v3.xlsm
  data/results/genre_labels_all.csv
  ailabel/results_claude.csv

出力:
  data/results/corpus_master.csv
  data/results/corpus_paths.csv
"""

import csv, re, os, sys, subprocess

try:
    import openpyxl
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl

from datetime import datetime

BASE        = r"D:\r8_strategy"
XLSM        = os.path.join(BASE, "data", "results", "audit_results_v3.xlsm")
GENRE_CSV   = os.path.join(BASE, "data", "results", "genre_labels_all.csv")
AILABEL_CSV = os.path.join(BASE, "ailabel", "results_claude.csv")
OUT_MASTER  = os.path.join(BASE, "data", "results", "corpus_master.csv")
OUT_PATHS   = os.path.join(BASE, "data", "results", "corpus_paths.csv")

MASTER_COLS = [
    "target",
    "cmi", "level",
    "human_label", "genre_label",
    "ailabel", "ailabel_model", "ailabel_condition",
    "ailabel_confidence", "ailabel_reason",
    "riskfactor_1", "riskfactor_2", "riskfactor_3",
    "remarks",
    "authority", "emotional", "logical", "statistical",
    "hype", "clickbait", "propaganda", "fear", "enemy_frame",
    "disclaimer_exploit", "anonymous_authority",
    "naked_number", "sexual_induction", "beauty_diet",
    "error", "timestamp",
]

PATHS_COLS = [
    "target", "primary_path", "retrieved_at",
    "content_hash", "status", "notes",
]

def normalize_target(raw):
    raw = str(raw).strip().replace("\uf8f0", "")
    if raw.startswith("http"):
        return raw
    if "\\" in raw or "/" in raw:
        normalized = raw.replace("\\", "/")
        parts = [p for p in normalized.split("/") if p]
        if "book" in [p.lower() for p in parts]:
            book_idx = next((i for i,p in enumerate(parts) if p.lower()=="book"), -1)
            if book_idx >= 0 and len(parts) > book_idx+2:
                folder = parts[book_idx+1]
                fname  = os.path.splitext(parts[-1])[0].replace("\uf8f0","")
                return f"book_{folder}_{fname}"
        fname = os.path.splitext(parts[-1])[0]
        fname = re.sub(r'_\d{8}$', '', fname)
        return fname
    if "." in raw and not raw.startswith("."):
        return os.path.splitext(raw)[0]
    return raw

def load_xlsm(path):
    print(f"  xlsm読み込み: {path}")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True, keep_vba=False)
    ws = wb["all_corpus"]
    data = {}
    for r in range(2, ws.max_row+1):
        raw_t = ws.cell(r, 2).value
        if not raw_t:
            continue
        norm = normalize_target(str(raw_t))
        data[norm] = {
            "raw_target":          str(raw_t).strip(),
            "timestamp":           ws.cell(r,  1).value,
            "cmi":                 ws.cell(r,  3).value,
            "level":               ws.cell(r,  4).value,
            "error":               ws.cell(r,  5).value,
            "authority":           ws.cell(r,  6).value,
            "emotional":           ws.cell(r,  7).value,
            "logical":             ws.cell(r,  8).value,
            "statistical":         ws.cell(r,  9).value,
            "hype":                ws.cell(r, 10).value,
            "clickbait":           ws.cell(r, 11).value,
            "propaganda":          ws.cell(r, 12).value,
            "fear":                ws.cell(r, 13).value,
            "enemy_frame":         ws.cell(r, 14).value,
            "disclaimer_exploit":  ws.cell(r, 15).value,
            "anonymous_authority": ws.cell(r, 16).value,
            "naked_number":        ws.cell(r, 17).value,
            "sexual_induction":    ws.cell(r, 18).value,
            "beauty_diet":         ws.cell(r, 19).value,
            "human_label":         ws.cell(r, 20).value,
            "riskfactor_1":        ws.cell(r, 21).value,
            "riskfactor_2":        ws.cell(r, 22).value,
            "riskfactor_3":        ws.cell(r, 23).value,
            "remarks":             ws.cell(r, 24).value,
        }
    wb.close()
    print(f"  -> {len(data)}件")
    return data

def load_genre(path):
    print(f"  genre_csv読み込み: {path}")
    data = {}
    if not os.path.exists(path):
        print(f"  WARNING ファイルなし: {path}")
        return data
    with open(path, encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            t = row["target"].strip()
            data[t] = {
                "genre_code": row.get("genre_code",""),
                "genre_name": row.get("genre_name",""),
            }
    print(f"  -> {len(data)}件")
    return data

def load_ailabel(path):
    print(f"  ailabel_csv読み込み: {path}")
    data = {}
    if not os.path.exists(path):
        print(f"  WARNING ファイルなし: {path}")
        return data
    with open(path, encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            raw_t = row.get("target","").strip()
            norm  = normalize_target(raw_t)
            data[norm] = {
                "ailabel":            row.get("ai_label",""),
                "ailabel_model":      row.get("model",""),
                "ailabel_condition":  row.get("primary_condition",""),
                "ailabel_confidence": row.get("confidence",""),
                "ailabel_reason":     row.get("reason",""),
            }
    print(f"  -> {len(data)}件")
    return data

def migrate():
    today = datetime.now().strftime("%Y-%m-%d")
    print("\n=== corpus_master移行スクリプト ===")
    print(f"実行日: {today}\n")

    xlsm_data    = load_xlsm(XLSM)
    genre_data   = load_genre(GENRE_CSV)
    ailabel_data = load_ailabel(AILABEL_CSV)

    master_rows     = []
    genre_matched   = 0
    ailabel_matched = 0

    for target, xd in xlsm_data.items():
        gd = genre_data.get(target, {})
        ad = ailabel_data.get(target, {})
        if gd: genre_matched   += 1
        if ad: ailabel_matched += 1

        genre_label = ""
        if gd.get("genre_code") and gd.get("genre_name"):
            genre_label = f"{gd['genre_code']}: {gd['genre_name']}"

        row = {
            "target":             target,
            "cmi":                xd.get("cmi",""),
            "level":              xd.get("level",""),
            "human_label":        xd.get("human_label",""),
            "genre_label":        genre_label,
            "ailabel":            ad.get("ailabel",""),
            "ailabel_model":      ad.get("ailabel_model",""),
            "ailabel_condition":  ad.get("ailabel_condition",""),
            "ailabel_confidence": ad.get("ailabel_confidence",""),
            "ailabel_reason":     ad.get("ailabel_reason",""),
            "riskfactor_1":       xd.get("riskfactor_1",""),
            "riskfactor_2":       xd.get("riskfactor_2",""),
            "riskfactor_3":       xd.get("riskfactor_3",""),
            "remarks":            xd.get("remarks",""),
            "authority":          xd.get("authority",""),
            "emotional":          xd.get("emotional",""),
            "logical":            xd.get("logical",""),
            "statistical":        xd.get("statistical",""),
            "hype":               xd.get("hype",""),
            "clickbait":          xd.get("clickbait",""),
            "propaganda":         xd.get("propaganda",""),
            "fear":               xd.get("fear",""),
            "enemy_frame":        xd.get("enemy_frame",""),
            "disclaimer_exploit": xd.get("disclaimer_exploit",""),
            "anonymous_authority":xd.get("anonymous_authority",""),
            "naked_number":       xd.get("naked_number",""),
            "sexual_induction":   xd.get("sexual_induction",""),
            "beauty_diet":        xd.get("beauty_diet",""),
            "error":              xd.get("error",""),
            "timestamp":          xd.get("timestamp",""),
        }
        master_rows.append(row)

    with open(OUT_MASTER, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=MASTER_COLS)
        writer.writeheader()
        writer.writerows(master_rows)

    print(f"\nOK corpus_master.csv: {len(master_rows)}件")
    print(f"   genre_label結合:  {genre_matched}件 / 空欄: {len(master_rows)-genre_matched}件")
    print(f"   ailabel結合:      {ailabel_matched}件 / 空欄: {len(master_rows)-ailabel_matched}件")

    paths_rows = []
    for target, xd in xlsm_data.items():
        paths_rows.append({
            "target":       target,
            "primary_path": xd["raw_target"],
            "retrieved_at": today,
            "content_hash": "",
            "status":       "active",
            "notes":        "migrated from audit_results_v3.xlsm",
        })

    with open(OUT_PATHS, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=PATHS_COLS)
        writer.writeheader()
        writer.writerows(paths_rows)

    print(f"\nOK corpus_paths.csv: {len(paths_rows)}件")

    # 整合性確認
    print("\n=== 整合性確認 ===")
    with open(OUT_MASTER, encoding="utf-8-sig") as f:
        check_rows = list(csv.DictReader(f))
    with open(OUT_PATHS, encoding="utf-8-sig") as f:
        check_paths = list(csv.DictReader(f))

    master_targets = {r["target"] for r in check_rows}
    paths_targets  = {r["target"] for r in check_paths}

    print(f"corpus_master件数:  {len(check_rows)}")
    print(f"corpus_paths件数:   {len(check_paths)}")
    print(f"target完全一致:     {len(master_targets & paths_targets)}件")

    cols_ok = list(check_rows[0].keys()) == MASTER_COLS
    print(f"列定義一致:         {'OK' if cols_ok else 'ERROR'} ({len(MASTER_COLS)}列)")

    print(f"\n出力先:")
    print(f"  {OUT_MASTER}")
    print(f"  {OUT_PATHS}")
    print("\n=== 移行完了 ===")

if __name__ == "__main__":
    migrate()
